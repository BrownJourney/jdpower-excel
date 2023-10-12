from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook
import openpyxl.utils.cell
import requests
from copy import copy

# JD Power API token
JDPOWER_AUTH_TOKEN = "<YOUR_AUTH_TOKEN>"
# Region ID for JD Power API request. See JD Power REST API guide to learn more
REGION_ID = "<YOUR_CURRENT_REGION>"

# @function | copy_style
# @args     | None
# @returns  | None
# @purpose  | copies cell style parameters to another cell
def copy_style(cell_style, new_cell):
    new_cell.font = copy(cell_style.font)
    new_cell.border = copy(cell_style.border)
    new_cell.fill = copy(cell_style.fill)
    new_cell.number_format = copy(cell_style.number_format)
    new_cell.protection = copy(cell_style.protection)
    new_cell.alignment = copy(cell_style.alignment)


class JDPowerExcel():
    def __init__(self):
        # File path input for user
        self.file_path = input("Enter file path:\n")
        self.generate_table()

    # @function | generate_table
    # @args     | None
    # @returns  | None
    # @purpose  | generates new Excel (xlsx) table with JD Power book values
    def generate_table(self):
        wb = None
        # Trying to load workbook from defined file path
        try:
            wb = load_workbook(self.file_path)
        # Catching case when file does not exist
        except FileNotFoundError:
            print("This file does not exists! Check defined path again!")
            return
        # Catching case when file extension is invalid
        except InvalidFileException:
            print("Invalid file extension!")
            return

        # Using active worksheet
        ws = wb.active
        # Defining columns for our future values. They are being set after last column in ascending order
        jd_clean_column = ws.max_column + 1
        jd_average_column = ws.max_column + 2
        jd_rough_column = ws.max_column + 3

        # Inserting these columns in worksheet
        ws.insert_cols(jd_clean_column)
        ws.insert_cols(jd_average_column)
        ws.insert_cols(jd_rough_column)

        # this value is used to store header offset
        reserve_headers = 0

        # vin_column and mileage_column define in which column is vin and mileage detected respectively
        vin_column = ''
        mileage_column = ''

        # iterating through rows in worksheet
        for row in ws.iter_rows():
            found_keywords = False
            reserve_headers = reserve_headers + 1
            for cell in row:
                # We are checking if row has at least some value
                if cell.value is not None:
                    val = cell.value.lower()
                    # Checking if value has vin or mileage (odometer) word in it.
                    found_vin = val.find("vin") != -1
                    found_mileage = ((val.find("odometer") != -1 or val.find("mileage") != -1)
                                     and val.find("unit") == -1)
                    if found_vin or found_mileage:
                        found_keywords = True

                    # Assigning columns in which VIN and Mileage values are stored
                    if found_vin:
                        vin_column = openpyxl.utils.cell.get_column_letter(cell.column)

                    if found_mileage:
                        mileage_column = openpyxl.utils.cell.get_column_letter(cell.column)

            if found_keywords:
                break

        # If VIN or Mileage is not present, abort process!
        if not (vin_column != '' and mileage_column != ''):
            print("This table does not contain VIN and Mileage to generate book values!")
            return

        # Defining JD Power headers
        jd_headers = {
            openpyxl.utils.cell.get_column_letter(jd_clean_column): {
                "name": "JD Trade Clean",
                "id": "adjustedcleantrade"
            },
            openpyxl.utils.cell.get_column_letter(jd_average_column): {
                "name": "JD Trade Average",
                "id": "adjustedaveragetrade"
            },
            openpyxl.utils.cell.get_column_letter(jd_rough_column): {
                "name": "JD Trade Rough",
                "id": "adjustedroughtrade"
            }
        }

        # Defining header and cell y offset to copy their styles for our JD Power columns
        header_offset_y = str(reserve_headers)
        cell_offset_y = str(reserve_headers + 1)

        # Getting header cell and default cell elements
        cell_style_header = ws[openpyxl.utils.cell.get_column_letter(ws.max_column) + header_offset_y]
        cell_style = ws[openpyxl.utils.cell.get_column_letter(ws.max_column) + cell_offset_y]

        # Assigning headers for new columnbs
        for key in jd_headers:
            ws[key + header_offset_y] = jd_headers[key]["name"]
            new_cell = ws[key + header_offset_y]

            # Copying header style for new header
            copy_style(cell_style_header, new_cell)
            # Setting good width for new column
            ws.column_dimensions[key].width = 20

        # initializing iterator
        i = 0

        for cell in ws[vin_column]:
            vin = cell.value
            mileage = ws[mileage_column][i].value

            i = i + 1

            # If we have not passed our header, do nothing
            if i <= reserve_headers:
                continue

            # URL of API request to the "defaultVehicleAndValuesByVin" method, which requires:
            # 1) period - defines which period book values should represent. Leave 0 if you want to get present values
            # 2) vin - Vehicle Identification Number
            # 3) region - region ID (from 0 to 10), which applies to the book values. For ex. California region is 10
            # 4) mileage - Vehicle odometer value
            url = ("https://cloud.jdpower.ai/data-api/UAT/valuationservices/valuation/defaultVehicleAndValuesByVin?"
                   "period=0&vin={0}&region={1}&mileage={2}")

            # Making GET-request for this url with vehicle vin and mileage
            response = requests.get(url.format(vin, REGION_ID, mileage), headers={
                # Authorization token
                "api-key": JDPOWER_AUTH_TOKEN,
                # Response type (you can also receive XML-format response if you want to)
                "accept": "application/json"
            })

            # Reading our response as JSON and converting it to the "dict"
            response = response.json()

            # "result" key appears only when API request was successful.
            if "result" not in response:
                print("API request failed!")
                continue

            # Returning first instance of result
            vehicle = response["result"][0]
            # If there is no instance, that means that JD Power has not found vehicle by declared VIN
            if not vehicle:
                print("Unable to find vehicle by this VIN (" + vin + ")!")
                break

            # Writing received values in JD Power book values cells
            for key in jd_headers:
                ws[key + str(i)] = "$" + str(vehicle[jd_headers[key]["id"]])
                copy_style(cell_style, ws[key + str(i)])

            print("Loaded book values for VIN={0} and MILEAGE={1}".format(vin, mileage))

        # Removing extension from pathname
        self.file_path.replace(".xlsx", "")
        # Saving generated table with same name and _bookvalues suffix
        wb.save(self.file_path + "_bookvalues.xlsx")
        print("Table is saved!")


JDPowerExcel()
